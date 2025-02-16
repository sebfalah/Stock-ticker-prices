import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

def main():
    # Get user input
    ticker_input = input("Enter the stock ticker: ").upper().strip()
    start_date = input("Enter the start date (YYYY-MM-DD): ").strip()
    end_date = input("Enter the end date (YYYY-MM-DD): ").strip()

    # Get ticker object from yfinance and retrieve stock info
    ticker = yf.Ticker(ticker_input)
    info = ticker.info

    # Retrieve the company name (using shortName or longName as available)
    stock_name = info.get('shortName') or info.get('longName') or "N/A"

    # Show the symbol and name for confirmation
    print(f"\nStock Symbol: {ticker_input}")
    print(f"Stock Name: {stock_name}")
    confirm = input("Is this correct? (Y/N): ").strip().lower()
    if confirm != 'y':
        print("Exiting the program.")
        return

    # Download historical data for the specified date range
    df = ticker.history(start=start_date, end=end_date)

    # Check if data is available
    if df.empty:
        print("No data found for the specified date range.")
        return

    # Reset the index to bring the Date into a column
    df = df.reset_index()

    # Remove timezone information from the Date column if present
    if isinstance(df['Date'].dtype, pd.DatetimeTZDtype):
        df['Date'] = df['Date'].dt.tz_localize(None)

    # Compute gain/loss using the provided formula: (Closing Price/Opening Price - 1)
    df['gain/loss'] = (df['Close'] / df['Open'] - 1)

    # Add stock symbol and stock name columns
    df['Stock Symbol'] = ticker_input
    df['Stock Name'] = stock_name

    # Rename columns to match the required output format
    df.rename(columns={'Open': 'Opening Price', 'Close': 'Closing Price'}, inplace=True)

    # Rearrange columns as desired
    final_df = df[['Stock Symbol', 'Stock Name', 'Date', 'Opening Price', 'Closing Price', 'gain/loss']]

    # Export the DataFrame to an Excel file
    file_name = f"{ticker_input}_stock_data.xlsx"
    final_df.to_excel(file_name, index=False)
    print(f"\nData has been saved to '{file_name}'.")

    # --- Create charts in a new sheet ---
    wb = load_workbook(file_name)
    # Assume the data is in the active sheet (Sheet1)
    ws = wb.active

    # Create a new sheet for the charts
    chart_sheet = wb.create_sheet("Chart")

    # Chart 1: Daily Percentage Change (gain/loss)
    chart1 = LineChart()
    chart1.title = "Daily Percentage Change"
    chart1.style = 13
    chart1.y_axis.title = "Percentage Change"
    chart1.x_axis.title = "Date"
    chart1.y_axis.number_format = '0.00%'
    chart1.legend = None  # Remove legend

    # Reference for gain/loss data: column F (6th column)
    data1 = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)
    chart1.add_data(data1, titles_from_data=True)

    # Set categories (x-axis) using the Date column: column C (3rd column)
    cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    chart1.set_categories(cats)

    # Add chart1 to the Chart sheet at cell A1
    chart_sheet.add_chart(chart1, "A1")

    # Chart 2: Daily Closing Prices
    chart2 = LineChart()
    chart2.title = "Daily Closing Price"
    chart2.style = 13
    chart2.y_axis.title = "Closing Price"
    chart2.x_axis.title = "Date"
    chart2.legend = None  # Remove legend

    # Reference for Closing Price data: column E (5th column)
    data2 = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    chart2.add_data(data2, titles_from_data=True)

    # Set categories (x-axis) using the Date column: column C (3rd column)
    chart2.set_categories(cats)

    # Add chart2 to the Chart sheet at cell A20 (adjust cell location as needed)
    chart_sheet.add_chart(chart2, "A20")

    # Save the workbook with the charts added
    wb.save(file_name)
    print(f"Charts have been added to the '{file_name}' in sheet 'Chart'.")

if __name__ == "__main__":
    main()
