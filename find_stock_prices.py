import yfinance as yf
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

def main():
    # Get user input for ticker symbol
    ticker_input = input("Enter the stock ticker: ").upper().strip()
    ticker = yf.Ticker(ticker_input)

    # Attempt to retrieve ticker info (e.g., company name) with retries
    max_retries = 3
    info = None
    for attempt in range(max_retries):
        try:
            # Use get_info() instead of .info
            info = ticker.get_info()
            # If .get_info() fails, you can try .fast_info
            # info = ticker.fast_info
            break
        except Exception as e:
            print(f"Error retrieving stock info (attempt {attempt+1} of {max_retries}): {e}")
            if attempt < max_retries - 1:
                print("Retrying in 5 seconds...")
                time.sleep(5)
            else:
                print("Failed to retrieve stock info after multiple attempts.")
    
    # If info retrieval failed, default to using the ticker symbol as the company name.
    if not info:
        print("Proceeding without company name information.")
        stock_name = ticker_input
    else:
        stock_name = info.get('shortName') or info.get('longName') or ticker_input

    # Display the ticker and company name for confirmation
    print(f"\nStock Symbol: {ticker_input}")
    print(f"Stock Name: {stock_name}")
    confirm = input("Is this correct? (Y/N): ").strip().lower()
    if confirm != 'y':
        print("Exiting the program.")
        return

    # Get date range from user
    start_date = input("Enter the start date (YYYY-MM-DD): ").strip()
    end_date = input("Enter the end date (YYYY-MM-DD): ").strip()
    
    # Fetch historical data
    df = ticker.history(start=start_date, end=end_date)
    # Alternatively: df = yf.download(ticker_input, start=start_date, end=end_date)

    if df.empty:
        print("No historical data found for the specified date range. Please check the dates or ticker symbol.")
        return

    df = df.reset_index()
    # Remove timezone if any
    if pd.api.types.is_datetime64tz_dtype(df['Date']):
        df['Date'] = df['Date'].dt.tz_localize(None)

    # Compute gain/loss
    df['gain/loss'] = (df['Close'] / df['Open'] - 1)

    df['Stock Symbol'] = ticker_input
    df['Stock Name'] = stock_name

    df.rename(columns={'Open': 'Opening Price', 'Close': 'Closing Price'}, inplace=True)
    final_df = df[['Stock Symbol', 'Stock Name', 'Date', 'Opening Price', 'Closing Price', 'gain/loss']]

    file_name = f"{ticker_input}_stock_data.xlsx"
    final_df.to_excel(file_name, index=False)
    print(f"\nData has been saved to '{file_name}'.")

    # --- Create charts in a new sheet ---
    wb = load_workbook(file_name)
    ws = wb.active

    chart_sheet = wb.create_sheet("Chart")

    # Chart 1: Daily Percentage Change (gain/loss)
    chart1 = LineChart()
    chart1.title = "Daily Percentage Change"
    chart1.style = 13
    chart1.y_axis.title = "Percentage Change"
    chart1.x_axis.title = "Date"
    chart1.y_axis.number_format = '0.00%'
    chart1.legend = None

    data1 = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)
    chart1.add_data(data1, titles_from_data=True)

    cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    chart1.set_categories(cats)
    chart_sheet.add_chart(chart1, "A1")

    # Chart 2: Daily Closing Prices
    chart2 = LineChart()
    chart2.title = "Daily Closing Price"
    chart2.style = 13
    chart2.y_axis.title = "Closing Price"
    chart2.x_axis.title = "Date"
    chart2.legend = None

    data2 = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart_sheet.add_chart(chart2, "A20")

    wb.save(file_name)
    print(f"Charts have been added to the '{file_name}' in sheet 'Chart'.")

if __name__ == "__main__":
    main()
